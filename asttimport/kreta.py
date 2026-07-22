import sys
import datetime as dt
import xml.etree.ElementTree as ET
from typing import Annotated

import openpyxl
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, BeforeValidator, Field
from pydantic_core.core_schema import ValidationInfo


TERM_DEFS: dict[str, tuple[dt.date, dt.date]] = {
    "YR": (dt.date(2026, 8, 31), dt.date(2027, 6, 25)),
    "F1": (dt.date(2026, 8, 31), dt.date(2027, 1, 22)),
    "F2": (dt.date(2027, 1, 25), dt.date(2027, 6, 25)),
}

EVERY_WEEK = "Minden héten"


def parse_comma_separated(v: str) -> list[str]:
    # print("SPLIT", v)
    return [item.strip() for item in v.split(",") if item.strip()]

def lookup_factory(key: str):
    def lookup(v, info: ValidationInfo):
        # print(f"LOOKUP {key}: '{v}'")
        if isinstance(v, list):
            return [getattr(info.context, key)[item] for item in v]
        elif isinstance(v, str):
            return getattr(info.context, key)[v]

        raise ValueError(f"Invalid value for {key}: '{v}'")

    return lookup


class XMLBaseModel(BaseModel):
    id: str
    name: str

    @property
    def kreta_name(self):
        return self.name

    @property
    def key(self):
        return self.id

class Day(XMLBaseModel):
    short: str
    days: Annotated[list[str], BeforeValidator(parse_comma_separated)]

    @property
    def key(self):
        return ",".join(self.days)

class Term(XMLBaseModel):
    short: str

class Subject(XMLBaseModel):
    short: str

    @property
    def kreta_name(self):
        if self.short.startswith("Digitális kultúra "):
            return "Digitális kultúra"
        elif self.short.startswith("Úszás ("):
            return "Úszás"
        elif self.short == "Testnevelés DU":
            return "Testnevelés"
        elif self.short.startswith("Angol nyelv "):
            return "Angol nyelv"
        return self.short


class Teacher(XMLBaseModel):
    firstname: str
    lastname: str
    short: str

    def __hash__(self):
        return hash(self.id)


class Classroom(XMLBaseModel):
    short: str


class Class(XMLBaseModel):
    short: str
    grade: Annotated[int | None, BeforeValidator(lambda x: int(x) if x else None)]


class Group(XMLBaseModel):
    entireclass: Annotated[bool, BeforeValidator(lambda x: x == "1")]
    class_: Annotated[Class, Field(validation_alias="classid"), BeforeValidator(lookup_factory("classes"))]
    is_assist: bool = Field(default=False)
    teachers: set[Teacher] = Field(default_factory=set)

    @property
    def kreta_name(self):
        assis_suffix = " - assziszt" if self.is_assist else ""
        teachers = '/'.join(sorted([f"{teacher.lastname[0]}{teacher.firstname[0]}" for teacher in self.teachers]))
        return f"{self.class_.name} - {self.name} - ({teachers}){assis_suffix}"

class Lesson(BaseModel):
    id: str
    classes: Annotated[list[Class], Field(validation_alias="classids"), BeforeValidator(lookup_factory("classes")), BeforeValidator(parse_comma_separated)]
    subject: Annotated[Subject, Field(validation_alias="subjectid"), BeforeValidator(lookup_factory("subjects"))]
    teachers: Annotated[list[Teacher], Field(validation_alias="teacherids"), BeforeValidator(lookup_factory("teachers")), BeforeValidator(parse_comma_separated)]
    term: Annotated[Term, Field(validation_alias="termsdefid"), BeforeValidator(lookup_factory("terms"))]
    groups: Annotated[list[Group], Field(validation_alias="groupids"), BeforeValidator(lookup_factory("groups")), BeforeValidator(parse_comma_separated)]

    @property
    def key(self):
        return self.id

class Card(BaseModel):
    lesson: Annotated[Lesson, Field(validation_alias="lessonid"), BeforeValidator(lookup_factory("lessons"))]
    classrooms: Annotated[list[Classroom], Field(validation_alias="classroomids"), BeforeValidator(lookup_factory("classrooms")), BeforeValidator(parse_comma_separated)]
    period: Annotated[int, BeforeValidator(int)]
    terms: str
    days: Annotated[list[Day], BeforeValidator(lookup_factory("days")), BeforeValidator(parse_comma_separated)]

    @property
    def key(self):
        return str(self)

class Timetable(BaseModel):
    days: Annotated[dict[str, Day], Day, "daysdefs/daysdef"] = Field(default_factory=dict)
    terms: Annotated[dict[str, Term], Term, "termsdefs/termsdef"] = Field(default_factory=dict)
    subjects: Annotated[dict[str, Subject], Subject, "subjects/subject"] = Field(default_factory=dict)
    teachers: Annotated[dict[str, Teacher], Teacher, "teachers/teacher"] = Field(default_factory=dict)
    classrooms: Annotated[dict[str, Classroom], Classroom, "classrooms/classroom"] = Field(default_factory=dict)
    classes: Annotated[dict[str, Class], Class, "classes/class"] = Field(default_factory=dict)
    groups: Annotated[dict[str, Group], Group, "groups/group"] = Field(default_factory=dict)
    lessons: Annotated[dict[str, Lesson], Lesson, "lessons/lesson"] = Field(default_factory=dict)
    cards: Annotated[dict[str, Card], Card, "cards/card"] = Field(default_factory=dict)


def load_xml(filename):
    tree = ET.parse(filename)
    return tree.getroot()


def parse_xml(xml: ET.Element):
    t = Timetable()
    for name, field in Timetable.model_fields.items():
        field_class, field_path = field.metadata
        objects = {}
        for item in xml.findall(field_path):
            obj = field_class.model_validate(item.attrib, context=t)
            objects[obj.key] = obj
        setattr(t, name, objects)

    return t

def extend_timetable_groups(timetable: Timetable):
    for lesson in timetable.lessons.values():
        assist_groups = []
        for group in lesson.groups:
            group.teachers.update(lesson.teachers)

            if len(lesson.teachers) > 1:
                assist_groups.append(group.model_copy(update={"is_assist": True}))
        if assist_groups:
            lesson.groups.extend(assist_groups)
            for assist_group in assist_groups:
                timetable.groups[assist_group.key] = assist_group

def write_sheet(workbook: openpyxl.Workbook, sheet_name: str, data: list[str] | set[str]):
    worksheet = workbook.create_sheet(sheet_name)
    for value in data:
        cell = WriteOnlyCell(worksheet, value)
        worksheet.append([cell])

    adjust_column_width(worksheet)

def write_cards(workbook: openpyxl.Workbook, timetable: Timetable):
    worksheet = workbook.create_sheet("Órarend érvényességi idővel")

    header_names = ["Óra érvényességének kezdete", "Óra érvényességének vége", "Hetirend", "Nap", "Óra (adott napon belül)", "Osztály", "Csoport", "Tantárgy", "Tanár", "Helyiség"]

    headers = []
    for header_name in header_names:
        cell = WriteOnlyCell(worksheet, header_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        headers.append(cell)
    worksheet.append(headers)

    for card in timetable.cards.values():
        lesson = card.lesson
        worksheet.append([
            TERM_DEFS[lesson.term.short][0],
            TERM_DEFS[lesson.term.short][1],
            EVERY_WEEK,
            ",".join(day.kreta_name for day in card.days),
            card.period,
            ",".join(class_.kreta_name for class_ in lesson.classes) if all(group.entireclass for group in lesson.groups) else None,
            ",".join(group.kreta_name for group in lesson.groups),
            lesson.subject.kreta_name,
            ",".join(teacher.kreta_name for teacher in lesson.teachers),
            ",".join(classroom.kreta_name for classroom in card.classrooms),
        ])

    adjust_column_width(worksheet)


def adjust_column_width(worksheet: Worksheet):
    for col in worksheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len

        adjusted_width = max_len + 3

        worksheet.column_dimensions[col_letter].width = adjusted_width

def write_excel(timetable: Timetable, filename: str):
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    write_cards(workbook, timetable)

    write_sheet(workbook, "Hetirend",[EVERY_WEEK] + [term.kreta_name for term in timetable.terms.values() if any(term.id == lesson.term.id for lesson in timetable.lessons.values())])
    write_sheet(workbook, "Nap", [day.kreta_name for day in timetable.days.values() if (len(day.days) == 1 and day.days[0].count("1") == 1)])
    write_sheet(workbook, "Osztály",
                [class_.kreta_name for class_ in timetable.classes.values()])

    write_sheet(workbook, "Csoport",
                [group.kreta_name for group in timetable.groups.values() if not group.entireclass])

    write_sheet(workbook, "Tantárgy",
                sorted({subject.kreta_name for subject in timetable.subjects.values()}))

    write_sheet(workbook, "Tanár",
                sorted({teacher.kreta_name for teacher in timetable.teachers.values()}))

    write_sheet(workbook, "Helyiség",
                sorted({classroom.kreta_name for classroom in timetable.classrooms.values()}))

    workbook.save(filename)


def main():
    xml = load_xml(sys.argv[1])

    timetable = parse_xml(xml)

    extend_timetable_groups(timetable)

    write_excel(timetable, sys.argv[2])


